import openpyxl

wb = openpyxl.load_workbook("ch09ex.xlsx")
sheet_names = wb.sheetnames
print(sheet_names)

work_sheet = wb["Python Ch09 Exercise"]

max_col = work_sheet.max_column
print("max column: ", max_col)
max_row = work_sheet.max_row
print("max row: ", max_row)

print("B2 셀의 값 : ", work_sheet["B7"].value)